import streamlit as st
import sys
from pathlib import Path
import tempfile
import os
from datetime import datetime
import pdfplumber

# Import the existing functions from bc_pdf_to_pivot
sys.path.append('.')
from bc_pdf_to_pivot import detect_format, parse_marjane, parse_lv, build_pivot

# Page configuration
st.set_page_config(
    page_title="PDF to Excel Pivot Converter",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS - Minimal Clean Design
st.markdown("""
<style>
    /* Header */
    .main-header {
        font-size: 2rem;
        font-weight: 600;
        color: #667eea;
        text-align: center;
        margin: 1rem 0 0.5rem 0;
    }
    
    .subtitle {
        text-align: center;
        color: #666;
        font-size: 0.95rem;
        margin-bottom: 2rem;
    }
    
    /* File List */
    .file-item {
        background: #f8f9fa;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        margin: 0.5rem 0;
        border-left: 3px solid #667eea;
        color: #333 !important;
    }
    
    .file-item div, .file-item span {
        color: #333 !important;
    }
    
    /* Button */
    .stButton > button {
        background: #667eea !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 0.8rem 2rem !important;
        font-weight: 500 !important;
    }
    
    .stButton > button:hover {
        background: #5a6fd6 !important;
    }
    
    /* Messages */
    .success-message {
        background: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #28a745;
    }
    
    .error-message {
        background: #f8d7da;
        color: #721c24;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #dc3545;
    }
    
    /* Download Button */
    .stDownloadButton > button {
        background: #28a745 !important;
        color: white !important;
        border-radius: 8px !important;
        border: none !important;
    }
    
    /* Hide Streamlit elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display: none;}
</style>
""", unsafe_allow_html=True)

def main():
    # Simple header
    st.markdown('<h1 class="main-header">📊 PDF to Excel Converter</h1>', unsafe_allow_html=True)
    st.markdown('<p class="subtitle">Convert BON DE COMMANDE PDFs to Excel pivot tables</p>', unsafe_allow_html=True)
    
    # File upload - clean and simple
    uploaded_files = st.file_uploader(
        "Drop PDF files here",
        type=['pdf'],
        accept_multiple_files=True
    )
    
    # Show files and process button
    if uploaded_files:
        st.markdown("### Selected Files")
        
        for file in uploaded_files:
            st.markdown(f'''
            <div class="file-item">
                📄 {file.name} ({(file.size/1024):.1f} KB)
            </div>
            ''', unsafe_allow_html=True)
        
        if st.button("Generate Excel", type="primary", use_container_width=True):
            process_uploaded_files(uploaded_files)
    
    # Sidebar - minimal
    with st.sidebar:
        st.markdown("### About")
        st.markdown("Convert PDF order forms to Excel pivot tables")
        st.markdown("**Formats:** Marjane, LV")
        st.markdown("---")
        st.markdown("v1.0")

def process_uploaded_files(uploaded_files):
    """Process uploaded PDF files and generate Excel files"""
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    results = []
    
    for i, uploaded_file in enumerate(uploaded_files):
        status_text.text(f"Processing {uploaded_file.name}...")
        progress_bar.progress((i + 0.5) / len(uploaded_files))
        
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(uploaded_file.getbuffer())
                tmp_path = tmp_file.name
            
            # Process the PDF
            result = process_single_pdf(tmp_path, uploaded_file.name)
            results.append(result)
            
            # Clean up temporary file
            os.unlink(tmp_path)
            
        except Exception as e:
            results.append({
                'filename': uploaded_file.name,
                'success': False,
                'error': str(e),
                'download_url': None
            })
        
        progress_bar.progress((i + 1) / len(uploaded_files))
    
    status_text.text("Processing complete!")
    
    # Display results
    display_results(results)

def process_single_pdf(pdf_path: str, filename: str) -> dict:
    """Process a single PDF file and return result dictionary"""
    try:
        # Detect format
        fmt = detect_format(pdf_path)
        
        # Debug: log detected format
        print(f"[DEBUG] Format detected for {filename}: {fmt}")
        
        # Parse data
        if fmt == "marjane":
            data, date_cmd, titre = parse_marjane(pdf_path)
        else:
            data, date_cmd, titre = parse_lv(pdf_path)
        
        # Debug: log extracted data count
        print(f"[DEBUG] Articles extracted: {len(data)}")
        
        if not data:
            return {
                'filename': filename,
                'success': False,
                'error': f"No data extracted. Format detected: {fmt.upper()}. The PDF may be scanned (image) or use an unsupported format.",
                'format': fmt
            }
        
        # Generate output filename
        stem = Path(filename).stem
        output_filename = f"pivot_{stem}.xlsx"
        
        # Create Excel file in memory
        excel_data = build_pivot_in_memory(data, titre, fmt)
        
        return {
            'filename': filename,
            'success': True,
            'format': fmt.upper(),
            'articles': len(data),
            'magasins': len(set(k for row in data.values() for k in row.keys() if k != 'libelle')),
            'date_cmd': date_cmd,
            'titre': titre,
            'excel_data': excel_data,
            'output_filename': output_filename,
            'download_url': None,
            'error': None
        }
        
    except Exception as e:
        import traceback
        print(f"[DEBUG] Error processing {filename}: {str(e)}")
        print(traceback.format_exc())
        return {
            'filename': filename,
            'success': False,
            'error': str(e),
            'download_url': None
        }

def build_pivot_in_memory(data: dict, titre: str, fmt: str) -> bytes:
    """Build Excel pivot table in memory and return bytes"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    if not data:
        raise ValueError("Aucune donnée à traiter")
    
    print(f"[DEBUG] Building pivot with {len(data)} articles")
    
    # Style constants (same as in bc_pdf_to_pivot.py)
    HEADER_BG = "1F4E79"
    HEADER_FG = "FFFFFF"
    TOTAL_BG = "D6E4F0"
    SUBHDR_BG = "BDD7EE"
    ALT_BG = "EBF3FB"
    
    def _border(style="thin"):
        s = Side(style=style)
        return Border(left=s, right=s, top=s, bottom=s)
    
    def _fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)
    
    def _font(bold=False, color="000000", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)
    
    # Collecter tous les magasins
    magasins = []
    seen = set()
    for ean, row in data.items():
        for k in row:
            if k != "libelle" and k not in seen:
                magasins.append(k)
                seen.add(k)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pivot BC"
    
    # ── Titre ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(magasins) + 1)
    title_cell = ws.cell(1, 1, titre)
    title_cell.font = Font(name="Arial", bold=True, color=HEADER_FG, size=12)
    title_cell.fill = _fill(HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    
    # ── En-têtes ──
    headers = ["EAN Article", "Libellé Article"] + magasins + ["TOTAL GÉNÉRAL"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font = _font(bold=True, color=HEADER_FG, size=10)
        c.fill = _fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[2].height = 30
    
    # ── Données ──
    EAN_col = 1
    LIB_col = 2
    first_mag_col = 3
    total_col = 2 + len(magasins) + 1
    
    rows_written = []
    for row_idx, (ean, row_data) in enumerate(data.items(), 3):
        alt = (row_idx % 2 == 0)
        bg = ALT_BG if alt else "FFFFFF"
        
        ws.cell(row_idx, EAN_col, str(ean)).border = _border()
        ws.cell(row_idx, EAN_col).font = _font()
        ws.cell(row_idx, EAN_col).fill = _fill(bg)
        
        ws.cell(row_idx, LIB_col, row_data.get("libelle", "")).border = _border()
        ws.cell(row_idx, LIB_col).font = _font()
        ws.cell(row_idx, LIB_col).fill = _fill(bg)
        
        for mag_idx, mag in enumerate(magasins):
            col = first_mag_col + mag_idx
            qty = row_data.get(mag, None)
            c = ws.cell(row_idx, col, qty)
            c.border = _border()
            c.font = _font()
            c.fill = _fill(bg)
            if qty is not None:
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="center")
        
        # Total ligne
        col_start = get_column_letter(first_mag_col)
        col_end = get_column_letter(first_mag_col + len(magasins) - 1)
        tc = ws.cell(row_idx, total_col, f"=SUM({col_start}{row_idx}:{col_end}{row_idx})")
        tc.border = _border()
        tc.font = _font(bold=True)
        tc.fill = _fill(TOTAL_BG)
        tc.number_format = "#,##0"
        tc.alignment = Alignment(horizontal="center")
        rows_written.append(row_idx)
    
    # ── Ligne TOTAL GÉNÉRAL ──
    if rows_written:
        total_row = rows_written[-1] + 1
        ws.cell(total_row, EAN_col, "TOTAL GÉNÉRAL").font = _font(bold=True)
        ws.cell(total_row, EAN_col).fill = _fill(TOTAL_BG)
        ws.cell(total_row, EAN_col).border = _border()
        ws.merge_cells(start_row=total_row, start_column=EAN_col,
                       end_row=total_row, end_column=LIB_col)
        
        for mag_idx in range(len(magasins)):
            col = first_mag_col + mag_idx
            col_letter = get_column_letter(col)
            r1, r2 = rows_written[0], rows_written[-1]
            c = ws.cell(total_row, col, f"=SUM({col_letter}{r1}:{col_letter}{r2})")
            c.font = _font(bold=True)
            c.fill = _fill(TOTAL_BG)
            c.border = _border()
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="center")
        
        # Total général de la dernière colonne
        tc_letter = get_column_letter(total_col)
        gt = ws.cell(total_row, total_col,
                     f"=SUM({tc_letter}{rows_written[0]}:{tc_letter}{rows_written[-1]})")
        gt.font = _font(bold=True)
        gt.fill = _fill(HEADER_BG)
        gt.font = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        gt.border = _border()
        gt.number_format = "#,##0"
        gt.alignment = Alignment(horizontal="center")
    
    # ── Largeurs colonnes ──
    ws.column_dimensions[get_column_letter(EAN_col)].width = 16
    ws.column_dimensions[get_column_letter(LIB_col)].width = 40
    for i in range(len(magasins)):
        ws.column_dimensions[get_column_letter(first_mag_col + i)].width = 18
    ws.column_dimensions[get_column_letter(total_col)].width = 16
    
    # ── Figer les volets ──
    ws.freeze_panes = ws.cell(3, first_mag_col)
    
    # Save to memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def display_results(results):
    """Display processing results"""
    st.markdown("### 📊 Processing Results")
    
    successful_results = [r for r in results if r['success']]
    failed_results = [r for r in results if not r['success']]
    
    if successful_results:
        # Success summary
        st.markdown(f'<div class="success-message">✅ Successfully processed {len(successful_results)} file(s)</div>', unsafe_allow_html=True)
        
        # Detailed results
        for result in successful_results:
            with st.expander(f"📄 {result['filename']} - {result['format']} Format"):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Articles", result['articles'])
                
                with col2:
                    st.metric("Stores", result['magasins'])
                
                with col3:
                    st.metric("Format", result['format'])
                
                if result['date_cmd']:
                    st.info(f"📅 Command Date: {result['date_cmd']}")
                
                # Download button
                st.download_button(
                    label=f"📥 Download {result['output_filename']}",
                    data=result['excel_data'],
                    file_name=result['output_filename'],
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
    
    if failed_results:
        # Failed results
        st.markdown(f'<div class="error-message">❌ Failed to process {len(failed_results)} file(s)</div>', unsafe_allow_html=True)
        
        for result in failed_results:
            st.error(f"**{result['filename']}**: {result['error']}")

# Footer
st.markdown("---")
st.markdown("### 💡 Tips")
st.markdown("""
- Upload multiple PDF files for batch processing
- The app automatically detects Marjane or LV formats
- Download individual Excel files for each processed PDF
- Check the detailed results for articles count and store information
""")

if __name__ == "__main__":
    main()
