
"""
bc_pdf_to_pivot.py — Extraction pivot BON DE COMMANDE (Marjane, Marjane-MEDIDIS & LV)
Usage: python bc_pdf_to_pivot.py <fichier.pdf> [output.xlsx]

Formats supportes:
  - medidis_livrea  : BON DE COMMANDE MEDIDIS — ordre colonnes: Commande par | Commande a | Livre a
                      Le magasin est dans la colonne "Livre a" (x > 350).
                      Ex: BC_MEDIDIS.pdf
  - medidis_cmdpar  : BON DE COMMANDE MEDIDIS — ordre colonnes: Commande par | Livre a | Commande a
                      Le magasin est dans la colonne "Commande par" (x < 180).
                      Ex: BC_MEDIDIS_OP_SAGA_190326.pdf
  - lv              : BON DE COMMANDE Hyper Marche LV / Hyper Sud
                      Meme structure que medidis_cmdpar mais avec colonnes "Code externe" + "Code EAN".
"""

import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# DÉTECTION DU FORMAT
#
# Tous les PDFs partagent la même grille x :
#   x ~95  : "Commande par"
#   x ~285 : "Livre a"  OU  "Commande a"  (selon le format)
#   x ~454 : "Commande a" OU "Livre a"
#
# medidis_livrea (BC_MEDIDIS) :
#   Commande par (x~95) | Commande a (x~276) | Livre a (x~463)
#   → MEDIDIS à x~208, magasin à x~386
#
# medidis_cmdpar (SAGA / LV) :
#   Commande par (x~95) | Livre a (x~285) | Commande a (x~454)
#   → magasin à x~30, MEDIDIS à x~386
#
# La distinction se fait sur la position du mot "Livrea" dans l'en-tête :
#   - x < 350  → format medidis_cmdpar (ou lv)
#   - x > 350  → format medidis_livrea
# ─────────────────────────────────────────────

def detect_format(pdf_path: str) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""
        words = pdf.pages[0].extract_words(x_tolerance=3, y_tolerance=3)

    upper = text.upper()

    # LV / Hyper Sud détecté sur le nom du magasin (les espaces peuvent être absents)
    upper_nospace = upper.replace(" ", "")
    if (
        "HYPER MARCHE LV" in upper
        or "HYPE MARCHE LV" in upper
        or "HYPER SUD" in upper
        or "HYPERMARCHELV" in upper_nospace
        or "HYPEMARCHELV" in upper_nospace
        or "HYPERSUD" in upper_nospace
    ):
        return "lv"

    # Trouver la position x du label "Livrea" dans les en-têtes de colonnes
    livrea_x = None
    for w in words:
        if w["text"].replace(" ", "").upper() == "LIVREA":
            livrea_x = w["x0"]
            break

    if livrea_x is not None:
        # "Livre a" loin à droite → magasin est à droite → format medidis_livrea
        if livrea_x > 350:
            return "medidis_livrea"
        # "Livre a" au centre → magasin est à gauche → format medidis_cmdpar
        return "medidis_cmdpar"

    # Fallback : si MARJANE présent, tenter medidis_livrea
    if "MARJANE" in upper:
        return "medidis_livrea"

    return "medidis_livrea"


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _get_rows(words, y_tolerance=3):
    rows = {}
    for w in words:
        y = round(w["top"] / y_tolerance) * y_tolerance
        rows.setdefault(y, []).append(w)
    return {y: sorted(ws, key=lambda w: w["x0"]) for y, ws in sorted(rows.items())}


def _normalize_magasin(raw: str) -> str:
    """Séparation des noms de magasin fusionnés : MARJANEBOUREGREG → MARJANE BOUREGREG"""
    s = re.sub(r"^(MARJANE)([A-Z])", r"\1 \2", raw, flags=re.I)
    return s.strip()


def _normalize_libelle(text: str) -> str:
    """Séparation des libellés fusionnés : CADREPHOTOZEYNA -> CADRE PHOTO ZEYNA"""
    s = text
    # Cadres photo
    s = re.sub(r"(CADRE)(PHOTO|MYLARD)", r"\1 \2", s, flags=re.I)
    s = re.sub(r"(PHOTO)(ZEYNA|LEA|GULIA|RITA|FLECHE)", r"\1 \2", s, flags=re.I)
    s = re.sub(r"(SILVER)(ROCK)", r"\1 \2", s, flags=re.I)
    # Plats cuisine
    s = re.sub(r"(PLATS?)(AFOUR|ATARTE)", r"\1 \2", s, flags=re.I)
    s = re.sub(r"(AFOUR|ATARTE)(RECT|ROND)", r"\1 \2", s, flags=re.I)
    s = re.sub(r"(RECT|ROND)(CERAM)", r"\1 \2", s, flags=re.I)
    s = re.sub(r"(CERAM)(PASSION)", r"\1 \2", s, flags=re.I)
    s = re.sub(r"(PASSION)(SMEG)", r"\1 \2", s, flags=re.I)
    # Divers
    s = re.sub(r"(SET\d*)(PLATS?)", r"\1 \2", s, flags=re.I)
    s = re.sub(r"(ROYAL)(VKB)", r"\1 \2", s, flags=re.I)
    return s.strip()


# ─────────────────────────────────────────────
# PARSEUR MEDIDIS — FORMAT "LIVRE A" (BC_MEDIDIS.pdf)
#
# Layout colonnes (positions x) :
#   x ~30-60  : EAN article
#   x ~111    : Libellé article (partie 1)
#   x ~208    : MEDIDIS (Commande a)
#   x ~233    : VL
#   x ~262    : No ligne  (13 chiffres, commence par 078…)
#   x ~341    : Type U.C. (PCB)
#   x ~386    : NOM DU MAGASIN (Livre a)
#   x ~403    : Quant en UC
#   x ~441    : UVC/UC
#   x ~474-496: Quant en UVC  ← valeur à extraire
#
# Ligne article sur 2 lignes y :
#   Ligne 1 : EAN | Libellé part1 | VL | No ligne | PCB | Quant UC | UVC | Quant UVC
#   Ligne 2 : Libellé part2 (suite du libellé, x~111)
# ─────────────────────────────────────────────

def parse_medidis_livrea(pdf_path: str) -> tuple[dict, str, str]:
    """
    Format BC_MEDIDIS :
    Ordre en-têtes : Commande par | Commande a | Livre a
    Magasin : colonne "Livre a" (x > 350).
    EAN : x < 60, Libellé : x ~79-160, Quant UVC : x >= 420.
    """
    data = {}
    date_cmd = ""

    EAN_RE       = re.compile(r"^\d{12,13}$")
    DATE_RE      = re.compile(r"(\d{2}/\d{2}/(?:20\d{2}|\d{2}))(?=[:\s]|\d{2}:|$)")
    NUM_RE       = re.compile(r"^\d+(\.\d+)?$")
    EAN_X_MAX    = 65     # EAN démarre à x~46
    LIBELLE_X    = 79     # libellé débute à x~111 mais on prend une marge
    LIBELLE_XMAX = 225    # limite droite du libellé (avant VL)
    NO_LIGNE_X   = 225    # No ligne à x~262 — au-delà de la zone libellé
    MAGASIN_X    = 350    # magasin (Livre a) à x~386
    QTY_UVC_X    = 420    # Quant en UVC à x~474-496

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue

            rows = _get_rows(words)
            sorted_ys = sorted(rows.keys())

            if not date_cmd:
                for ws in rows.values():
                    row_str = " ".join(w["text"] for w in ws)
                    m = DATE_RE.search(row_str)
                    if m:
                        date_cmd = m.group(1)
                        break

            # Magasin : ligne où MEDIDIS apparaît à x ~208, le magasin est à x > MAGASIN_X
            magasin = ""
            for ws in rows.values():
                texts = [w["text"].upper() for w in ws]
                if "MEDIDIS" in texts:
                    mag_words = [w for w in ws if w["x0"] > MAGASIN_X]
                    if mag_words:
                        raw = " ".join(w["text"] for w in mag_words)
                        magasin = _normalize_magasin(raw.split()[0])
                    break

            if not magasin:
                continue

            for idx, y in enumerate(sorted_ys):
                ws = rows[y]
                if not ws:
                    continue
                first = ws[0]
                if first["x0"] > EAN_X_MAX:
                    continue
                if not EAN_RE.match(first["text"]):
                    continue

                ean = first["text"]

                # Libellé : mots entre LIBELLE_X et LIBELLE_XMAX, non numériques
                lib_part1 = " ".join(
                    w["text"] for w in ws
                    if LIBELLE_X - 5 <= w["x0"] < LIBELLE_XMAX and not NUM_RE.match(w["text"])
                )

                # Libellé suite sur ligne suivante
                lib_part2 = ""
                if idx + 1 < len(sorted_ys):
                    next_ws = rows[sorted_ys[idx + 1]]
                    if next_ws and next_ws[0]["x0"] > EAN_X_MAX - 5:
                        lib_part2 = " ".join(
                            w["text"] for w in next_ws
                            if LIBELLE_X - 5 <= w["x0"] < LIBELLE_XMAX
                        )

                libelle = _normalize_libelle((lib_part1 + " " + lib_part2).strip())

                # Quant en UVC : dernier token numérique à droite (x >= QTY_UVC_X)
                uvc_candidates = [
                    w["text"] for w in ws if w["x0"] >= QTY_UVC_X and NUM_RE.match(w["text"])
                ]
                if not uvc_candidates:
                    continue
                qty = float(uvc_candidates[-1])

                if ean not in data:
                    data[ean] = {"libelle": libelle}
                data[ean][magasin] = data[ean].get(magasin, 0) + qty

    titre = "BON DE COMMANDE — MEDIDIS / MARJANE HOLDING"
    if date_cmd:
        titre += f" — {date_cmd}"
    return data, date_cmd, titre


# ─────────────────────────────────────────────
# PARSEUR MEDIDIS — FORMAT "COMMANDE PAR" (SAGA / standard)
#
# Layout colonnes (positions x) :
#   x ~30     : EAN article  ET  NOM DU MAGASIN (Commande par)
#   x ~79     : Libellé article
#   x ~160    : VL
#   x ~170    : No ligne (13 chiffres)
#   x ~226    : Type U.C. (PCB)
#   x ~262    : Quant en UC
#   x ~293    : UVC/UC
#   x ~323    : Quant en UVC  ← valeur à extraire
#   x ~356    : No. opération spéciale (OSSAGA3, optionnel)
#
# Magasin : ligne où MEDIDIS apparaît à x~386,
#            le magasin est dans les mots à x < 180.
#
# Ligne article sur 2 lignes y :
#   Ligne 1 : EAN | Libellé part1 | VL | No ligne | PCB | Qt UC | UVC/UC | Qt UVC [| OSSAGA3]
#   Ligne 2 : Libellé part2 (x~79)
# ─────────────────────────────────────────────

def parse_medidis_cmdpar(pdf_path: str) -> tuple[dict, str, str]:
    """
    Format SAGA / standard Marjane :
    Ordre en-têtes : Commande par | Livre a | Commande a
    Magasin : colonne "Commande par" (x < 180).
    EAN : x < 60, Libellé : x ~79-160, Quant UVC : x >= 310.
    """
    data = {}
    date_cmd = ""

    EAN_RE       = re.compile(r"^\d{12,13}$")
    DATE_RE      = re.compile(r"(\d{2}/\d{2}/(?:20\d{2}|\d{2}))(?=[:\s]|\d{2}:|$)")
    NUM_RE       = re.compile(r"^\d+(\.\d+)?$")
    EAN_X_MAX    = 60     # EAN à x~30
    LIBELLE_X    = 79     # libellé à x~79
    NO_LIGNE_X   = 160    # No ligne à x~170 — limite droite du libellé
    MEDIDIS_X    = 350    # MEDIDIS à x~386
    MAGASIN_X    = 180    # magasin à x < 180
    QTY_UVC_X    = 310    # Quant en UVC à x~323

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue

            rows = _get_rows(words)
            sorted_ys = sorted(rows.keys())

            if not date_cmd:
                for ws in rows.values():
                    row_str = " ".join(w["text"] for w in ws)
                    m = DATE_RE.search(row_str)
                    if m and "/" in m.group(1):
                        date_cmd = m.group(1)
                        break

            # Magasin : ligne où MEDIDIS apparaît à x > MEDIDIS_X
            magasin = ""
            for ws in rows.values():
                texts_upper = [w["text"].upper() for w in ws]
                if "MEDIDIS" in texts_upper:
                    store_words = [w for w in ws if w["x0"] < MAGASIN_X]
                    if store_words:
                        magasin = _normalize_magasin(store_words[0]["text"])
                    break

            if not magasin:
                continue

            for idx, y in enumerate(sorted_ys):
                ws = rows[y]
                if not ws:
                    continue
                first = ws[0]
                if first["x0"] > EAN_X_MAX:
                    continue
                if not EAN_RE.match(first["text"]):
                    continue

                ean = first["text"]

                # Libellé partie 1
                lib_part1 = " ".join(
                    w["text"] for w in ws
                    if LIBELLE_X - 5 <= w["x0"] < NO_LIGNE_X and not NUM_RE.match(w["text"])
                )

                # Libellé suite (ligne suivante)
                lib_part2 = ""
                if idx + 1 < len(sorted_ys):
                    next_ws = rows[sorted_ys[idx + 1]]
                    if next_ws and next_ws[0]["x0"] > EAN_X_MAX - 5:
                        lib_part2 = " ".join(
                            w["text"] for w in next_ws
                            if LIBELLE_X - 5 <= w["x0"] < NO_LIGNE_X
                        )

                libelle = _normalize_libelle((lib_part1 + " " + lib_part2).strip())

                # Quant en UVC
                uvc_candidates = [
                    w["text"] for w in ws if w["x0"] >= QTY_UVC_X and NUM_RE.match(w["text"])
                ]
                if not uvc_candidates:
                    continue
                qty = float(uvc_candidates[0])

                if ean not in data:
                    data[ean] = {"libelle": libelle}
                data[ean][magasin] = data[ean].get(magasin, 0) + qty

    titre = "BON DE COMMANDE — MEDIDIS / MARJANE HOLDING"
    if date_cmd:
        titre += f" — {date_cmd}"
    return data, date_cmd, titre


# ─────────────────────────────────────────────
# PARSEUR LV (Hyper Marché LV / Hyper Sud)
#
# Layout colonnes (positions x) :
#   x ~30     : NOM DU MAGASIN (Commande par, x < 180)
#   x ~45     : Code externe
#   x ~95     : Code EAN (13 chiffres)
#   x ~215    : Libellé article
#   x ~386    : MEDIDIS (Commande a)
#   x ~457    : UVC/UC
#   x ~498    : Quant en UC  ← valeur à extraire (= Quant en UVC car UVC/UC=1)
#
# Une ligne article = une ligne y (pas de continuation libellé sur 2 lignes).
# ─────────────────────────────────────────────

def parse_lv(pdf_path: str) -> tuple[dict, str, str]:
    """
    Format Hyper Marché LV / Hyper Sud.
    Ordre en-têtes : Commande par | Livre a | Commande a
    Magasin : colonne "Commande par" (x < 180).
    EAN : colonne "Code EAN" (x ~95), Quant UC : x >= 480.
    """
    data = {}
    date_cmd = ""

    EAN_RE    = re.compile(r"^\d{13}$")
    DATE_RE   = re.compile(r"(\d{2}/\d{2}/(?:20\d{2}|\d{2}))(?=[:\s]|\d{2}:|$)")
    NUM_RE    = re.compile(r"^\d+(\.\d+)?$")

    # Seuils x
    EAN_X_MIN   = 80      # Code EAN à x~95
    EAN_X_MAX   = 140
    LIBELLE_X   = 140     # libellé à x~215 (large marge gauche)
    LIBELLE_XMAX= 330     # limite droite libellé
    MAGASIN_X   = 180     # magasin (Commande par) à x < 180
    MEDIDIS_X   = 350     # MEDIDIS à x~386
    QTY_UC_X    = 480     # Quant en UC à x~498-510

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue

            rows = _get_rows(words)
            sorted_ys = sorted(rows.keys())

            if not date_cmd:
                for ws in rows.values():
                    row_str = " ".join(w["text"] for w in ws)
                    m = DATE_RE.search(row_str)
                    if m:
                        date_cmd = m.group(1)
                        break

            # Magasin : ligne où MEDIDIS apparaît à x > MEDIDIS_X
            magasin = ""
            for ws in rows.values():
                texts_upper = [w["text"].upper() for w in ws]
                if "MEDIDIS" in texts_upper:
                    store_words = [w for w in ws if w["x0"] < MAGASIN_X]
                    if store_words:
                        magasin = store_words[0]["text"].strip()
                    break

            if not magasin:
                continue

            for idx, y in enumerate(sorted_ys):
                ws = rows[y]
                if not ws:
                    continue

                # EAN dans la plage x prévue
                ean_words = [
                    w for w in ws
                    if EAN_X_MIN <= w["x0"] <= EAN_X_MAX and EAN_RE.match(w["text"])
                ]
                if not ean_words:
                    continue

                ean = ean_words[0]["text"]

                # Libellé partie 1 : mots entre LIBELLE_X et LIBELLE_XMAX
                lib_part1 = " ".join(
                    w["text"] for w in ws
                    if LIBELLE_X <= w["x0"] < LIBELLE_XMAX
                ).strip()

                # Libellé partie 2 : ligne suivante au même x (dimensions, ex: 39X22X8CM)
                lib_part2 = ""
                if idx + 1 < len(sorted_ys):
                    next_ws = rows[sorted_ys[idx + 1]]
                    # Ligne de continuation : pas d'EAN, a du texte à x~LIBELLE_X
                    has_ean = any(
                        EAN_X_MIN <= w["x0"] <= EAN_X_MAX and EAN_RE.match(w["text"])
                        for w in next_ws
                    )
                    if not has_ean and next_ws:
                        lib_part2 = " ".join(
                            w["text"] for w in next_ws
                            if LIBELLE_X <= w["x0"] < LIBELLE_XMAX
                        ).strip()

                libelle = _normalize_libelle((lib_part1 + " " + lib_part2).strip())

                # Quant en UC : dernier token numérique à x >= QTY_UC_X
                qty_candidates = [
                    w["text"] for w in ws if w["x0"] >= QTY_UC_X and NUM_RE.match(w["text"])
                ]
                if not qty_candidates:
                    continue
                qty = float(qty_candidates[-1])

                if ean not in data:
                    data[ean] = {"libelle": libelle}
                data[ean][magasin] = data[ean].get(magasin, 0) + qty

    titre = "BON DE COMMANDE — MEDIDIS / LV"
    if date_cmd:
        titre += f" — {date_cmd}"
    return data, date_cmd, titre


# ─────────────────────────────────────────────
# CONSTRUCTION DU PIVOT EXCEL
# ─────────────────────────────────────────────

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
TOTAL_BG  = "D6E4F0"
ALT_BG    = "EBF3FB"


def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)


def build_pivot(data: dict, titre: str, output_path: str, fmt: str):
    if not data:
        print("Aucune donnee extraite.")
        return

    magasins = []
    seen = set()
    for row in data.values():
        for k in row:
            if k != "libelle" and k not in seen:
                magasins.append(k)
                seen.add(k)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot BC"

    EAN_col       = 1
    LIB_col       = 2
    first_mag_col = 3
    total_col     = 2 + len(magasins) + 1

    # Ligne 1 : Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)
    c = ws.cell(1, 1, titre)
    c.font = Font(name="Arial", bold=True, color=HEADER_FG, size=12)
    c.fill = _fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 21.95

    # Ligne 2 : En-têtes
    for col, label in [(EAN_col, "EAN Article"), (LIB_col, "Libelle Article")]:
        c = ws.cell(2, col, label)
        c.font = _font(bold=True, color=HEADER_FG)
        c.fill = _fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()

    for i, mag in enumerate(magasins):
        c = ws.cell(2, first_mag_col + i, mag)
        c.font = _font(bold=True, color=HEADER_FG)
        c.fill = _fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
        c.border = _border()

    c = ws.cell(2, total_col, "TOTAL GENERAL")
    c.font = _font(bold=True, color=HEADER_FG)
    c.fill = _fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    ws.row_dimensions[2].height = 165.0

    # Lignes de données
    rows_written = []
    for row_idx, (ean, row_data) in enumerate(data.items(), 3):
        bg = ALT_BG if (row_idx % 2 == 0) else "FFFFFF"

        c = ws.cell(row_idx, EAN_col, str(ean))
        c.font = _font()
        c.fill = _fill(bg)
        c.border = _border()

        c = ws.cell(row_idx, LIB_col, row_data.get("libelle", ""))
        c.font = _font()
        c.fill = _fill(bg)
        c.border = _border()

        for i, mag in enumerate(magasins):
            qty = row_data.get(mag, None)
            c = ws.cell(row_idx, first_mag_col + i, qty)
            c.font = _font()
            c.fill = _fill(bg)
            c.border = _border()
            if qty is not None:
                c.number_format = "# ###"
                c.alignment = Alignment(horizontal="center")

        col_s = get_column_letter(first_mag_col)
        col_e = get_column_letter(first_mag_col + len(magasins) - 1)
        c = ws.cell(row_idx, total_col, f"=SUM({col_s}{row_idx}:{col_e}{row_idx})")
        c.font = _font(bold=True)
        c.fill = _fill(TOTAL_BG)
        c.border = _border()
        c.number_format = "# ###"
        c.alignment = Alignment(horizontal="center")
        rows_written.append(row_idx)

    # Ligne TOTAL GENERAL
    if rows_written:
        tr = rows_written[-1] + 1
        r1, r2 = rows_written[0], rows_written[-1]

        ws.merge_cells(start_row=tr, start_column=EAN_col, end_row=tr, end_column=LIB_col)
        c = ws.cell(tr, EAN_col, "TOTAL GENERAL")
        c.font = _font(bold=True)
        c.fill = _fill(TOTAL_BG)
        c.border = _border()

        for i in range(len(magasins)):
            col = first_mag_col + i
            cl = get_column_letter(col)
            c = ws.cell(tr, col, f"=SUM({cl}{r1}:{cl}{r2})")
            c.font = _font(bold=True)
            c.fill = _fill(TOTAL_BG)
            c.border = _border()
            c.number_format = "# ###"
            c.alignment = Alignment(horizontal="center")

        tcl = get_column_letter(total_col)
        c = ws.cell(tr, total_col, f"=SUM({tcl}{r1}:{tcl}{r2})")
        c.font = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        c.fill = _fill(HEADER_BG)
        c.border = _border()
        c.number_format = "# ###"
        c.alignment = Alignment(horizontal="center")

    # Largeurs colonnes
    ws.column_dimensions[get_column_letter(EAN_col)].width = 14.14
    ws.column_dimensions[get_column_letter(LIB_col)].width = 31.29
    for i in range(len(magasins)):
        ws.column_dimensions[get_column_letter(first_mag_col + i)].width = 3.29
    ws.column_dimensions[get_column_letter(total_col)].width = 13.0

    ws.freeze_panes = ws.cell(3, first_mag_col)
    wb.save(output_path)
    print(
        f"Pivot genere : {output_path} | Format: {fmt.upper()} "
        f"| Articles: {len(data)} | Magasins: {len(magasins)}"
    )


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def process_pdf(pdf_path: str, output_path: str):
    if not Path(pdf_path).exists():
        print(f"Fichier introuvable : {pdf_path}")
        return

    fmt = detect_format(pdf_path)
    print(f"Format detecte : {fmt.upper()}")

    if fmt == "medidis_livrea":
        data, date_cmd, titre = parse_medidis_livrea(pdf_path)
    elif fmt == "medidis_cmdpar":
        data, date_cmd, titre = parse_medidis_cmdpar(pdf_path)
    elif fmt == "lv":
        data, date_cmd, titre = parse_lv(pdf_path)
    else:
        # Fallback conservatif
        data, date_cmd, titre = parse_medidis_livrea(pdf_path)

    build_pivot(data, titre, output_path, fmt)


def main():
    if len(sys.argv) >= 2:
        pdf_path = sys.argv[1]
        output_path = (
            sys.argv[2] if len(sys.argv) >= 3 else f"pivot_{Path(pdf_path).stem}.xlsx"
        )
        process_pdf(pdf_path, output_path)
        return

    pdf_files = list(Path(".").glob("*.pdf"))
    if not pdf_files:
        print(
            "Aucun fichier PDF trouve. "
            "Usage: python bc_pdf_to_pivot.py <fichier.pdf> [output.xlsx]"
        )
        sys.exit(1)

    for pdf_file in pdf_files:
        print(f"\n{'='*50}\nTraitement : {pdf_file.name}")
        process_pdf(str(pdf_file), f"pivot_{pdf_file.stem}.xlsx")


if __name__ == "__main__":
    main()